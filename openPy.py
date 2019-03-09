import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# creating an instance of file object
def process_workbook(filename):

    wb = xl.load_workbook(filename)
    # sheets could be duynamic too
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        corrected_price = sheet.cell(row, 3).value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(
        sheet,
        min_row=2,
        max_row=sheet.max_row,
        min_col=4,
        max_col=4
    )

    chart = BarChart()
    chart.add_data(values)
    # injecting column can be dynamic too
    sheet.add_chart(chart, 'E2')

    wb.save(filename)


